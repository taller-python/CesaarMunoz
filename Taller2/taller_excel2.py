""" Taller lectura archivos excel y escritura archivos texto """
import openpyxl
wb = openpyxl.load_workbook('operaciones.xlsx')
file = open("taller_excel_salida.txt", "w")
# sumas
hoja = wb['SUMA']
file.write('procesando '+hoja.title+'\n')
# realizar las operaciones
for r in range(2, hoja.max_row+1):
    try:
        file.write(str(int(hoja.cell(row=r, column=1).value)
                       + int(hoja.cell(row=r, column=2).value))+'\n')
    except ValueError:
        file.write('Error'+'\n')

# RESTAS
hoja = wb['RESTA']
file.write('procesando '+hoja.title+'\n')
# realizar las operaciones
for r in range(2, hoja.max_row+1):
    try:
        file.write(str(int(hoja.cell(row=r, column=1).value)
                       - int(hoja.cell(row=r, column=2).value))+'\n')
    except ValueError:
        file.write('Error'+'\n')

# MULTIPLICACION
hoja = wb['MULTIPLICACIÓN']
file.write('procesando '+hoja.title+'\n')
# realizar las operaciones
for r in range(2, hoja.max_row+1):
    try:
        file.write(str(int(hoja.cell(row=r, column=1).value)
                       * int(hoja.cell(row=r, column=2).value))+'\n')
    except ValueError:
        file.write('Error'+'\n')

# DIVISION
hoja = wb['DIVISIÓN']
file.write('procesando '+ hoja.title+'\n')
# realizar las operaciones
for r in range(2, hoja.max_row+1):
    try:
        file.write(str(int(hoja.cell(row=r, column=1).value)
                       / int(hoja.cell(row=r, column=2).value))+'\n')
    except ValueError:
        file.write('Error'+'\n')
    except ZeroDivisionError:
        file.write('Error'+'\n')
file.close()
