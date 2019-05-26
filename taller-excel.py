""" Taller lectura archivos excel y escritura archivos texto """
import openpyxl
wb = openpyxl.load_workbook('operaciones.xlsx')
# sumas
hoja = wb['SUMA']
print('procesando '+hoja.title)
# realizar las operaciones
for r in range(2, hoja.max_row+1):
    try:
        print(int(hoja.cell(row=r, column=1).value) + int(hoja.cell(row=r, column=2).value))
    except:
        print('Error')
    
# RESTAS
hoja = wb['RESTA']
print('procesando '+hoja.title)
# realizar las operaciones
for r in range(2, hoja.max_row+1):
    try:
        print(int(hoja.cell(row=r, column=1).value) - int(hoja.cell(row=r, column=2).value))
    except:
        print('Error')

# MULTIPLICACION
hoja = wb['MULTIPLICACIÓN']
print('procesando '+hoja.title)
# realizar las operaciones
for r in range(2, hoja.max_row+1):
    try:
        print(int(hoja.cell(row=r, column=1).value) * int(hoja.cell(row=r, column=2).value))
    except:
        print('Error')

# DIVISION
hoja = wb['DIVISIÓN']
print('procesando '+ hoja.title)
# realizar las operaciones
for r in range(2, hoja.max_row+1):
    try:
        print(int(hoja.cell(row=r, column=1).value) / int(hoja.cell(row=r, column=2).value))
    except:
        print('Error')
